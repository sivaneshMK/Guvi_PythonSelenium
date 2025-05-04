from openpyxl import Workbook
wb = Workbook()
#sheet = wb.create_sheet("Guvi")
sheet = wb.create_sheet()
sheet['A1'] = "Indhumathi"
wb.save("C://Users//sivan//PycharmProjects//Guvi_PythonSelenium//FileHandling//GuviData.xlsx")
